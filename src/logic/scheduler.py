from src.data.schedule_entry import ScheduleEntry
from src.logic.validator import validate_schedule
from openpyxl.utils import get_column_letter

def get_time_slots():
    MONDAY_SLOTS = [
        (0, "08:30", "09:15"),
        (1, "09:20", "10:50"),
        (2, "11:10", "12:40"),
        (3, "12:50", "14:20"),
        (4, "14:35", "16:05"),
        (5, "16:10", "17:40"),
        (6, "17:45", "19:15"),
    ]
    OTHER_DAYS_SLOTS = [
        (1, "08:30", "10:00"),
        (2, "10:10", "11:40"),
        (3, "12:00", "13:30"),
        (4, "14:00", "15:30"),
        (5, "15:40", "17:10"),
        (6, "17:15", "18:45"),
    ]
    slots = {"Понедельник": MONDAY_SLOTS}
    for day in ["Вторник", "Среда", "Четверг", "Пятница", "Суббота"]:
        slots[day] = OTHER_DAYS_SLOTS
    return slots

def generate_schedule(teachers, rooms, groups, time_slots):
    # Словарь для быстрого доступа по ID аудиторий.
    room_dict = {room.id: room for room in rooms}
    # occupancy[day][pair] – контролирует занятость: какие преподаватели и кабинеты уже задействованы.
    occupancy = {day: {} for day in time_slots}
    # schedule[day] – список ScheduleEntry для данного дня.
    schedule = {day: [] for day in time_slots}
    # group_slots[day] – для каждой группы (по group_id) список назначенных пар (для последовательности и не более 4 пар).
    group_slots = {day: {} for day in time_slots}

    for day, slots in time_slots.items():
        for pair, start, end in slots:
            occupancy[day][pair] = {"teachers": set(), "rooms": set()}
        # Обработка понедельника, слот 0 (ЧКР): для каждой группы создаём запись без преподавателя.
        if day == "Понедельник":
            for pair, start, end in slots:
                if pair == 0:
                    for group in groups:
                        chosen_room = None
                        for room in rooms:
                            if not room.available:
                                continue
                            if day in room.absences and pair in room.absences[day]:
                                continue
                            if room.id in occupancy[day][pair]["rooms"]:
                                continue
                            chosen_room = room
                            break
                        if chosen_room:
                            entry = ScheduleEntry(day, pair, start, end, None, chosen_room, group.name, "ЧКР")
                            schedule[day].append(entry)
                            occupancy[day][pair]["rooms"].add(chosen_room.id)
                            group_slots[day].setdefault(group.id, []).append(pair)
                        else:
                            entry = ScheduleEntry(day, pair, start, end, None, None, group.name, "ЧКР (Дистанционно)")
                            schedule[day].append(entry)
                            group_slots[day].setdefault(group.id, []).append(pair)
        # Обработка остальных слотов (исключая для понедельника слот 0).
        for teacher in teachers:
            if not teacher.available:
                continue
            for assignment in teacher.assignments:
                group_id = assignment.get("group_id")
                group_name = next((g.name for g in groups if g.id == group_id), f"Группа {group_id}")
                if group_id in group_slots[day] and len(group_slots[day][group_id]) >= 4:
                    continue  # Ограничение: не более 4 пар для группы
                slots_sorted = sorted(slots, key=lambda x: x[0])
                scheduled = False
                desired_slot = None
                if group_id in group_slots[day]:
                    desired_slot = max(group_slots[day][group_id]) + 1
                available_slots = [ (pair, start, end) for pair, start, end in slots_sorted
                                    if not (day=="Понедельник" and pair==0)
                                    and pair not in teacher.absences.get(day, []) ]
                if desired_slot is not None:
                    candidate_slots = [s for s in available_slots if s[0] == desired_slot]
                    if candidate_slots:
                        pair, start, end = candidate_slots[0]
                        chosen_room = None
                        if teacher.preferred_rooms:
                            for room_id in teacher.preferred_rooms:
                                if room_id not in room_dict:
                                    continue
                                room = room_dict[room_id]
                                if day in room.absences and pair in room.absences[day]:
                                    continue
                                if room.id in occupancy[day][pair]["rooms"]:
                                    continue
                                chosen_room = room
                                break
                        if not chosen_room:
                            for room in rooms:
                                if not room.available:
                                    continue
                                if day in room.absences and pair in room.absences[day]:
                                    continue
                                if room.id in occupancy[day][pair]["rooms"]:
                                    continue
                                chosen_room = room
                                break
                        if chosen_room:
                            subject_str = f"{assignment.get('subject', 'Без предмета')} ({teacher.name}), {group_name}"
                            entry = ScheduleEntry(day, pair, start, end, teacher, chosen_room, group_name, subject_str)
                            schedule[day].append(entry)
                            occupancy[day][pair]["teachers"].add(teacher.id)
                            occupancy[day][pair]["rooms"].add(chosen_room.id)
                            group_slots[day].setdefault(group_id, []).append(pair)
                            scheduled = True
                    if scheduled:
                        continue
                for pair, start, end in available_slots:
                    if group_id in group_slots[day]:
                        last_assigned = max(group_slots[day][group_id])
                        if pair != last_assigned + 1:
                            continue
                    if teacher.id in occupancy[day][pair]["teachers"]:
                        continue
                    chosen_room = None
                    if teacher.preferred_rooms:
                        for room_id in teacher.preferred_rooms:
                            if room_id not in room_dict:
                                continue
                            room = room_dict[room_id]
                            if day in room.absences and pair in room.absences[day]:
                                continue
                            if room.id in occupancy[day][pair]["rooms"]:
                                continue
                            chosen_room = room
                            break
                    if not chosen_room:
                        for room in rooms:
                            if not room.available:
                                continue
                            if day in room.absences and pair in room.absences[day]:
                                continue
                            if room.id in occupancy[day][pair]["rooms"]:
                                continue
                            chosen_room = room
                            break
                    if chosen_room:
                        subject_str = f"{assignment.get('subject', 'Без предмета')} ({teacher.name}), {group_name}"
                        entry = ScheduleEntry(day, pair, start, end, teacher, chosen_room, group_name, subject_str)
                        schedule[day].append(entry)
                        occupancy[day][pair]["teachers"].add(teacher.id)
                        occupancy[day][pair]["rooms"].add(chosen_room.id)
                        group_slots[day].setdefault(group_id, []).append(pair)
                        scheduled = True
                        break
                # Если ни один слот не подходит, запись не создается (дистанционный режим применяем только при явном отсутствии преподавателя).
        schedule[day].sort(key=lambda e: e.period)
    return schedule

def export_schedule_to_excel(schedule, groups, filename="schedule.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook()
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    # Определяем даты для дней (пример).
    day_date = {
        "Понедельник": "03.03.2025",
        "Вторник": "04.03.2025",
        "Среда": "05.03.2025",
        "Четверг": "06.03.2025",
        "Пятница": "07.03.2025",
        "Суббота": "08.03.2025"
    }
    # Группируем расписание по группе (наименование).
    group_schedules = {}
    for day, entries in schedule.items():
        for entry in entries:
            grp = entry.group if entry.group else "Не указана"
            group_schedules.setdefault(grp, {}).setdefault(day, []).append(entry)
    # Для каждой группы создаём отдельный лист.
    for grp, day_entries in group_schedules.items():
        ws = wb.create_sheet(title=str(grp)[:31])
        current_row = 1
        for day in ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]:
            if day not in day_entries:
                continue
            # Заголовок дня с датой.
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws.cell(row=current_row, column=1, value=f"{day} - {day_date.get(day, '')}")
            cell.font = header_font
            cell.alignment = header_alignment
            current_row += 1
            # Заголовок таблицы для дня.
            headers = ["Пара", "Предмет", "Преподаватель", "Аудитория"]
            for col_index, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_index, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
            current_row += 1
            # Сортируем записи по номеру пары.
            day_entries[day].sort(key=lambda e: e.period)
            for entry in day_entries[day]:
                teacher_name = entry.teacher.name if entry.teacher is not None else "Не назначен"
                room_name = entry.room.name if entry.room is not None else "Не назначена"
                ws.cell(row=current_row, column=1, value=entry.period)
                ws.cell(row=current_row, column=2, value=entry.subject)
                ws.cell(row=current_row, column=3, value=teacher_name)
                ws.cell(row=current_row, column=4, value=room_name)
                current_row += 1
            # Пустая строка между днями.
            current_row += 1
        # Автоматическая настройка ширины столбцов.
        for i, col in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(i)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
    # Удаляем стандартный лист, если он пуст.
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        del wb["Sheet"]
    wb.save(filename)
    print(f"Расписание экспортировано в файл {filename}.")

def generate_and_validate_schedule(teachers, rooms, groups):
    time_slots = get_time_slots()
    schedule = generate_schedule(teachers, rooms, groups, time_slots)
    errors = validate_schedule(schedule)
    return schedule, errors
